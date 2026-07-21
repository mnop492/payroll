import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_standard_template(calc_month, brand_code='century_field', db_path='payroll.db', output_folder='history'):
    """
    產生包含當月所有 Data 與 Excel 運算公式的標準化模板
    """
    os.makedirs(output_folder, exist_ok=True)
    filename = f"{calc_month.replace('-', '')}_{brand_code}_Standard_Template.xlsx"
    filepath = os.path.join(output_folder, filename)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # 1. 撈取【員工主檔與月度微調】(決定時薪、津貼、佣金率與 MPF 狀態)
    emp_df = pd.read_sql_query("""
        SELECT e.nick_name, e.full_name, e.salary_type, e.monthly_salary, 
               e.hourly_rate as default_hr, e.allowance as default_allowance, 
               e.commission_rate as default_comm, e.require_mpf, e.mpf_start_month,
               mr.hourly_rate as monthly_hr, mr.allowance as monthly_allowance, 
               mr.commission_rate as monthly_comm
        FROM Employees e
        LEFT JOIN MonthlyRates mr ON e.nick_name = mr.nick_name AND mr.payroll_month=? AND mr.brand_code=?
        WHERE e.brand_code=?
    """, conn, params=(calc_month, brand_code, brand_code))
    
    emp_dict = {}
    for _, row in emp_df.iterrows():
        name = row['nick_name']
        emp_dict[name] = {
            'full_name': row['full_name'] or '',
            'salary_type': row['salary_type'] or 'hourly',
            'eff_hr': row['monthly_hr'] if pd.notna(row['monthly_hr']) else row['default_hr'],
            'eff_allowance': row['monthly_allowance'] if pd.notna(row['monthly_allowance']) else row['default_allowance'],
            'eff_comm': row['monthly_comm'] if pd.notna(row['monthly_comm']) else row['default_comm'],
            'monthly_salary': row['monthly_salary'] or 0,
            'require_mpf': row['require_mpf'],
            'mpf_start_month': row['mpf_start_month']
        }

    # 2. 撈取【每日考勤】
    daily_df = pd.read_sql_query("""
        SELECT work_date, nick_name, location, in_time, out_time, 
               actual_hours, ot_hours, roster_in, roster_out, normal_hours 
        FROM DailyAttendance WHERE brand_code=? AND payroll_month=?
    """, conn, params=(brand_code, calc_month))

    # 3. 撈取【當月既有銷售明細】
    sales_df = pd.read_sql_query("""
        SELECT s.date, s.promoter_name, s.location, s.model, s.quantity, s.price,
               p.product_line, p.product_category, p.commission_rate as prod_comm, 
               sc.rate as spec_comm
        FROM Sales s
        LEFT JOIN Products p ON s.model = p.model AND p.brand_code = s.brand_code
        LEFT JOIN SpecialCommissions sc ON s.model = sc.model AND sc.brand_code = s.brand_code 
             AND sc.start_month <= ? AND sc.end_month >= ?
        WHERE s.brand_code=? AND s.payroll_month=?
    """, conn, params=(calc_month, calc_month, brand_code, calc_month))

    # 4. 撈取【月度微調】(報銷、微調、出勤獎)
    adj_df = pd.read_sql_query("""
        SELECT nick_name, location, expenses, adjustment, attendance_bonus 
        FROM Attendance WHERE brand_code=? AND payroll_month=?
    """, conn, params=(brand_code, calc_month))
    
    # 🌟 5. 新增：撈取【當月完整的產品清單、特佣規則與所有 DB 欄位】
    products_df = pd.read_sql_query("""
        SELECT p.model, p.product_line, p.product_category, p.commission_rate as default_comm,
               sc.rate as special_comm, sc.start_month, sc.end_month,
               p.id as product_id, p.brand_code
        FROM Products p
        LEFT JOIN SpecialCommissions sc ON p.model = sc.model 
             AND sc.brand_code = p.brand_code
             AND sc.start_month <= ? AND sc.end_month >= ?
        WHERE p.brand_code=?
    """, conn, params=(calc_month, calc_month, brand_code))
    
    # ✅ 所有的資料庫查詢都完成後，再統一關閉連線
    conn.close()

    # ==========================================
    # 開始建立 Excel Workbook
    # ==========================================
    wb = openpyxl.Workbook()

    # --- 樣式定義 ---
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    def apply_header_style(ws, row_idx=1):
        for cell in ws[row_idx]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align            

    # ------------------------------------------
    # 分頁 1: 工作表1 (考勤明細)
    # ------------------------------------------
    ws_ts = wb.active
    ws_ts.title = "工作表1"
    ts_headers = ["Date", "Nick Name", "Location", "Name", "In-Time", "Out-Time", "Hours", "OT Hours", "In-Time2", "Out-Time2", "Normal working hours"]
    ws_ts.append(ts_headers)
    apply_header_style(ws_ts)

    for _, row in daily_df.iterrows():
        ws_ts.append([
            row['work_date'], row['nick_name'], row['location'], row['nick_name'], 
            row['in_time'], row['out_time'], row['actual_hours'], row['ot_hours'], 
            row['roster_in'], row['roster_out'], row['normal_hours']
        ])

    # 🌟 --- 新增分頁: Product List (包含所有 DB 欄位與當月生效佣金) --- 🌟
    ws_prod = wb.create_sheet(title="Product List")
    prod_headers = [
        "Model", "Product Line", "Category", "Default Comm Rate", 
        "Special Comm Rate", "Special Start", "Special End", 
        "Effective Comm Rate", "Product ID", "Brand Code"
    ]
    ws_prod.append(prod_headers)
    apply_header_style(ws_prod)
    
    for _, p in products_df.iterrows():
        # 決定當月生效的佣金率：有特佣就用特佣，沒有就用預設
        eff_rate = p['special_comm'] if pd.notna(p['special_comm']) else p['default_comm']
        
        ws_prod.append([
            p['model'], 
            p['product_line'] if p['product_line'] else '', 
            p['product_category'] if p['product_category'] else '', 
            p['default_comm'],
            p['special_comm'] if pd.notna(p['special_comm']) else '', 
            p['start_month'] if pd.notna(p['start_month']) else '', 
            p['end_month'] if pd.notna(p['end_month']) else '', 
            eff_rate,       # H欄 (第8欄): 最終計算用的生效佣金率
            p['product_id'],             
        ])

    # 🌟 --- 分頁 2: editable-sales-form (銷售明細與 VLOOKUP 自動查表) --- 🌟
    ws_sales = wb.create_sheet(title="editable-sales-form")
    # 將 Category 加入到欄位中
    sales_headers = ["Shop", "Location", "Promoter", "Month", "Date", "Model", "Product Line", "Category", "Input", "Quantity", "Price", "Comm Rate", "Comm Amount"]
    ws_sales.append(sales_headers)
    apply_header_style(ws_sales)

    current_sales_row = 2
    
    # A. 寫入資料庫「既有的」當月銷售紀錄
    for _, row in sales_df.iterrows():
        name = row['promoter_name']
        emp_info = emp_dict.get(name, {})
        # 決定套用哪層佣金 (優先級: 特佣 > 產品預設 > 員工專屬 > 系統保底 3%)
        eff_comm_rate = row['spec_comm'] if pd.notna(row['spec_comm']) else \
                        row['prod_comm'] if pd.notna(row['prod_comm']) else \
                        emp_info.get('eff_comm', 0.03)

        ws_sales.append([
            row['location'], row['location'], name, calc_month, row['date'], 
            row['model'], row['product_line'], row['product_category'], '', row['quantity'], row['price'], 
            eff_comm_rate,
            f"=J{current_sales_row}*K{current_sales_row}*L{current_sales_row}" # 數量(J)*單價(K)*佣金率(L)
        ])
        current_sales_row += 1

    # B. 預留 100 行「智慧空白列」供 User 離線新增單據
    for _ in range(100):
        # 只要 F 欄(Model) 有填字，就去 Product List 查表，查不到給空值或預設值
        vlookup_line = f'=IF(F{current_sales_row}="","", IFERROR(VLOOKUP(F{current_sales_row}, \'Product List\'!A:D, 2, FALSE), ""))'
        vlookup_cat  = f'=IF(F{current_sales_row}="","", IFERROR(VLOOKUP(F{current_sales_row}, \'Product List\'!A:D, 3, FALSE), ""))'
        vlookup_rate = f'=IF(F{current_sales_row}="","", IFERROR(VLOOKUP(F{current_sales_row}, \'Product List\'!A:D, 4, FALSE), 0.03))'
        
        # Comm Amount 公式：數量(J) * 單價(K) * 佣金率(L)
        calc_formula = f'=IF(F{current_sales_row}="","", J{current_sales_row}*K{current_sales_row}*L{current_sales_row})'
        
        ws_sales.append([
            "", "", "", calc_month, "", 
            "",            # F: Model
            vlookup_line,  # G: Product Line
            vlookup_cat,   # H: Category
            "",            # I: Input
            1,             # J: Quantity
            0,             # K: Price
            vlookup_rate,  # L: Comm Rate
            calc_formula   # M: Comm Amount
        ])
        current_sales_row += 1

    # B. 預留 100 行「智慧空白列」供 User 離線新增單據
    # 利用 Excel 的 VLOOKUP 自動去 Product List 尋找型號對應的佣金率！
    for _ in range(100):
        # K欄 (Comm Rate) 公式：如果 F欄(Model) 有填字，就去 Product List 查表，找不到預設給 0.03
        vlookup_formula = f'=IF(F{current_sales_row}="","", IFERROR(VLOOKUP(F{current_sales_row}, \'Product List\'!A:D, 4, FALSE), 0.03))'
        
        # L欄 (Comm Amount) 公式：數量(I) * 單價(J) * 佣金率(K)
        calc_formula = f'=IF(F{current_sales_row}="","", I{current_sales_row}*J{current_sales_row}*K{current_sales_row})'
        
        ws_sales.append(["", "", "", calc_month, "", "", "", "", 1, 0, vlookup_formula, calc_formula])
        current_sales_row += 1

    # ------------------------------------------
    # 分頁 3: Total (薪資總結 - 全公式化)
    # ------------------------------------------
    ws_total = wb.create_sheet(title="Total")
    
    # 為了排版美觀，前6行留空 (符合歷史系統的讀取習慣 header=6)
    for _ in range(6):
        ws_total.append([])
        
    total_headers = [
        "Name", "Shop", "Days", "Hours", "Hourly Rate", "酬  金", 
        "Basic Comm", "Allowance", "Expenses", "Adjustment", 
        "Attendance", "Total", "MPF", "Net Pay"
    ]
    ws_total.append(total_headers)
    
    # ------------------------------------------
    # 分頁 3: Total (薪資總結 - 全公式化)
    # ------------------------------------------
    ws_total = wb.create_sheet(title="Total")
    
    # 為了排版美觀，前6行留空 (符合歷史系統的讀取習慣 header=6)
    for _ in range(6):
        ws_total.append([])
        
    # 🌟 補上 Full Name 與 Sales Qty
    total_headers = [
        "Name", "Full Name", "Shop", "Days", "Hours", "Hourly Rate", "酬  金", 
        "Sales Qty", "Basic Comm", "Allowance", "Expenses", "Adjustment", 
        "Attendance", "Total", "MPF", "Net Pay"
    ]
    ws_total.append(total_headers)
    
    # 樣式：Total 頁面標頭
    for col_idx in range(1, len(total_headers) + 1):
        cell = ws_total.cell(row=7, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 找出當月有出勤或有銷售的獨特 (Nick Name, Location) 組合
    unique_pairs = set()
    for _, row in daily_df.iterrows():
        unique_pairs.add((row['nick_name'], row['location']))
    for _, row in sales_df.iterrows():
        unique_pairs.add((row['promoter_name'], row['location']))
    
    start_row = 8
    for i, (name, loc) in enumerate(sorted(list(unique_pairs))):
        current_row = start_row + i
        emp_info = emp_dict.get(name, {})
        
        # 取得手動微調項目
        adj_record = adj_df[(adj_df['nick_name'] == name) & (adj_df['location'] == loc)]
        exp = adj_record['expenses'].values[0] if not adj_record.empty else 0
        adj = adj_record['adjustment'].values[0] if not adj_record.empty else 0
        bonus = adj_record['attendance_bonus'].values[0] if not adj_record.empty else 0
        
        # 決定時薪與每日津貼，加上嚴格的空值防呆
        hr_rate = emp_info.get('eff_hr')
        hr_rate = float(hr_rate) if pd.notna(hr_rate) and hr_rate is not None else 0
        
        daily_allowance = emp_info.get('eff_allowance')
        daily_allowance = float(daily_allowance) if pd.notna(daily_allowance) and daily_allowance is not None else 0
        
        full_name = emp_info.get('full_name', '')

        # A: Name, B: Full Name, C: Shop
        ws_total.cell(row=current_row, column=1, value=name)
        ws_total.cell(row=current_row, column=2, value=full_name)
        ws_total.cell(row=current_row, column=3, value=loc)
        
        # D: Days (使用 COUNTIFS 統計工作表1的打卡天數，地點改對齊 C 欄)
        ws_total.cell(row=current_row, column=4, value=f"=COUNTIFS('工作表1'!$B:$B, A{current_row}, '工作表1'!$C:$C, C{current_row})")
        
        # E: Hours (直接加總工作表1的 K欄 Normal working hours)
        ws_total.cell(row=current_row, column=5, value=f"=SUMIFS('工作表1'!$K:$K, '工作表1'!$B:$B, A{current_row}, '工作表1'!$C:$C, C{current_row})")
        
        # F: Hourly Rate
        ws_total.cell(row=current_row, column=6, value=hr_rate)
        
        # G: 酬金 (Basic Pay)
        if emp_info.get('salary_type') == 'monthly':
            ws_total.cell(row=current_row, column=7, value=emp_info.get('monthly_salary', 0)) # 月薪制直接寫入死薪水
        else:
            ws_total.cell(row=current_row, column=7, value=f"=E{current_row}*F{current_row}") # 時薪制 = 工時(E)*時薪(F)
            
        # 🌟 H: Sales Qty (使用 SUMIFS 統計 editable-sales-form 的 J 欄 Quantity)
        ws_total.cell(row=current_row, column=8, value=f"=SUMIFS('editable-sales-form'!$J:$J, 'editable-sales-form'!$C:$C, A{current_row}, 'editable-sales-form'!$B:$B, C{current_row})")

        # I: Basic Comm (使用 SUMIFS 統計 editable-sales-form 的 M 欄 Comm Amount)
        ws_total.cell(row=current_row, column=9, value=f"=SUMIFS('editable-sales-form'!$M:$M, 'editable-sales-form'!$C:$C, A{current_row}, 'editable-sales-form'!$B:$B, C{current_row})")
        
        # J: Allowance (天數(D) * 每日津貼)
        ws_total.cell(row=current_row, column=10, value=f"=D{current_row}*{daily_allowance}")
        
        # K, L, M: Expenses, Adjustment, Attendance
        ws_total.cell(row=current_row, column=11, value=exp)
        ws_total.cell(row=current_row, column=12, value=adj)
        ws_total.cell(row=current_row, column=13, value=bonus)
        
        # N: Total Gross Pay (酬金 + 佣金 + 津貼 + 報銷 + 微調 + 勤工獎) 
        # ⚠️ 注意公式避開 H 欄的數量：SUM(G, I 到 M)
        ws_total.cell(row=current_row, column=14, value=f"=SUM(G{current_row}, I{current_row}:M{current_row})")
        
        # O: MPF
        mpf_start = emp_info.get('mpf_start_month')
        require_mpf = emp_info.get('require_mpf', 1)
        if not require_mpf or (mpf_start and calc_month < mpf_start):
            ws_total.cell(row=current_row, column=15, value=0)
        else:
            ws_total.cell(row=current_row, column=15, value=f"=IF(N{current_row}>=7100, MIN(ROUND(N{current_row}*0.05, 0), 1500), 0)")
            
        # P: Net Pay (Total - MPF) - 使用 ROUND 確保數學準確至小數點 2 位，並強制顯示 2 位小數
        net_pay_cell = ws_total.cell(row=current_row, column=16, value=f"=ROUND(N{current_row}-O{current_row}, 2)")
        net_pay_cell.number_format = '#,##0.00'

    # 移除預設的空 Sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(filepath)
    return filepath